import re
import openpyxl
import os
import win32com.client


def extract_infos_from_hwp(hwp_file):
    infos = []
    hwp = None
    try:
        hwp = win32com.client.Dispatch("HWPFrame.HwpObject")
        hwp.RegisterModule("FilePathCheckDLL", "SecurityModule")
        hwp.Open(hwp_file)
        hwp.InitScan()
        text = hwp.GetText()

        while True:
            if text[0] == 0:
                hwp.ReleaseScan()
                break
            else:
                found_emails = re.findall(
                    r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', text)

                for email in found_emails:
                    infos.append((os.path.basename(hwp_file),
                                  hwp.key_indicator()[3], "이메일",  email))

    except Exception as e:
        print(f"한글파일 에러발생 : {e}")
    finally:
        if hwp:
            try:
                hwp.ReleaseControl()
            except:
                pass

    return infos


def save_infos_to_excel(emails, excel_file):
    if os.path.exists(excel_file):
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["연번", "파일명", "페이지번호", "유형", "내용"])
    start_no = sheet.max_row if sheet.cell(
        row=1, column=1).value == "No." else 0

    for i, email in enumerate(emails, start=start_no + 1):
        sheet.append([i] + list(email))

    workbook.save(excel_file)


def processing_folder(folder_path, excel_file):
    all_emails = []

    for filename in os.listdir(folder_path):
        if filename.lower().endswith('.hwp') or filename.lower().endswith('.hwpx'):
            hwp_file_path = os.path.join(folder_path, filename)
            emails = extract_infos_from_hwp(hwp_file_path)
            all_emails.extend(emails)

    save_infos_to_excel(all_emails, excel_file)


if __name__ == "__main__":
    folder_path = input("개인정보를 추출할 파일이 있는 폴더 경로를 입력하세요: ")
    excel_file = input(
        "엑셀파일 경로를 입력하세요(확장자포함. 파일이 존재하지 않을 경우 새로 생성): ")
    processing_folder(folder_path, excel_file)
    print(f"{excel_file}에 개인정보목록이 생성되었습니다.")
