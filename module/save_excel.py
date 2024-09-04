import openpyxl


import os


def save_infos_to_excel(infos, excel_file):
    if os.path.exists(excel_file):
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["연번", "파일명", "페이지번호", "유형", "내용"])
    start_no = sheet.max_row if sheet.cell(
        row=1, column=1).value == "No." else 0

    for i, info in enumerate(infos, start=start_no + 1):
        sheet.append([i] + list(info))

    workbook.save(excel_file)
