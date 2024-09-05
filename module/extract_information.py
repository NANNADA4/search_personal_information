"""
동일한 정규표현식을 사용하여 각 파일에 맞는 방법으로 개인정보를 추출합니다
"""

import os
import re
import win32com.client as win32
import fitz
from openpyxl import load_workbook


PATTERN_EMAILS = r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}'
PATTERN_JUMINS = r'\d{2}[01]\d[0123]\d- [1-4]\d{6}'
PATTERN_CREDIT_NUMS = r'\b\d{4}-\d{4}-\d{4}-\d{4}\b'
PATTERN_CELLPHONE_NUMS = r'\b(010-\d{4}-\d{4}|01[16789]-\d{3,4}-\d{4})\b'
PATTERN_DRIVER_NUMS = r'\d{2}-\d{2}-\d{6}-\d{2}'


PATTERNS = {
    '이메일': PATTERN_EMAILS,
    '주민등록번호': PATTERN_JUMINS,
    '신용카드번호': PATTERN_CREDIT_NUMS,
    '휴대전화번호': PATTERN_CELLPHONE_NUMS,
    '운전면허번호': PATTERN_DRIVER_NUMS
}


def _extract_personal_information(folder_path, file, text=None, page_num=None, sheet_name=None, error=None):
    """정규표현식으로 개인정보를 추출하여 리스트로 return합니다"""
    infos = []

    blank = str(os.path.basename(folder_path)).find(' ')
    under_bar = str(os.path.basename(folder_path)).find('_')
    if blank != -1 and under_bar != -1:
        cmt = str(os.path.basename(folder_path))[blank+1:under_bar]
    elif blank != -1 and under_bar == -1:
        cmt = str(os.path.basename(folder_path))[blank+1:]
    else:
        cmt = str(os.path.basename(folder_path))
    relative_path = os.path.relpath(file, os.path.dirname(folder_path))

    if text is None:
        infos.append(
            (cmt, relative_path.split(os.sep)[1],
             os.path.basename(file), None, None, None, error))

        return infos

    for info_type, pattern in PATTERNS.items():
        matches = re.findall(pattern, text)
        for match in matches:
            # 위원회, 피감기관, 파일명, 페이지수, 개인정보 종류, 개인정보 검색 결과, 에러
            if page_num is not None:
                infos.append(
                    (cmt, relative_path.split(os.sep)[1],
                     os.path.basename(file), page_num + 1, info_type, match, None))
            elif sheet_name is not None:
                infos.append(
                    (cmt, relative_path.split(os.sep)[1],
                     os.path.basename(file), sheet_name, info_type, match, None))
            else:
                infos.append((cmt, relative_path.split(
                    os.sep)[1], os.path.basename(file), None, info_type, match, None))

    return infos


def processing_pdf(folder_path, pdf_file):
    """pdf파일을 처리후, pdf_infos에 모든 결과를 리스트로 저장하여 return합니다"""
    try:
        doc = fitz.open(pdf_file)
        pdf_infos = []

        for page_num in range(len(doc)):
            page = doc.load_page(page_num)
            text = page.get_text()
            pdf_infos.extend(_extract_personal_information(folder_path,
                                                           pdf_file, text=text, page_num=page_num))
    except Exception as e:  # pylint: disable=W0703
        error_log = str(e)
        pdf_infos.extend(
            _extract_personal_information(folder_path, pdf_file, error=error_log))
        print(pdf_file, e)

    return pdf_infos


def processing_hwp(folder_path, hwp_file):
    """hwp 파일을 처리 후, hwp_infos에 모든 결과를 리스트로 저장하여 반환합니다"""
    hwp_infos = []
    hwp = None

    try:
        hwp = win32.gencache.EnsureDispatch("HWPFrame.HwpObject")
        hwp.RegisterModule("FilePathCheckDLL", "SecurityModule")
        hwp.Open(hwp_file)
        hwp.InitScan()

        while True:
            state, text = hwp.GetText()
            if state in [0, 1]:
                break
            hwp_infos.extend(
                _extract_personal_information(folder_path, hwp_file, text=text))

    except Exception as e:  # pylint: disable=W0703
        error_log = str(e)
        hwp_infos.extend(
            _extract_personal_information(folder_path, hwp_file, error=error_log))
        print(hwp_file, e)

    finally:
        if hwp:
            hwp.ReleaseScan()
            hwp.Quit()

    return hwp_infos


def processing_excel(folder_path, excel_file):
    """엑셀 파일을 처리 후, excel_infos에 모든 결과를 리스트로 저장하여 반환합니다"""
    excel_infos = []
    try:
        wb = load_workbook(excel_file)

        for sheet in wb.sheetnames:
            ws = wb[sheet]
            text = ""
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if cell:
                        text += str(cell) + " "
            excel_infos.extend(_extract_personal_information(folder_path,
                                                             excel_file, text=text, sheet_name=sheet))
    except Exception as e:  # pylint: disable=W0703
        error_log = str(e)
        excel_infos.extend(
            _extract_personal_information(folder_path, excel_file, error=error_log))
        print(excel_file, e)

    return excel_infos
