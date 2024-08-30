import fitz
import re
import openpyxl
import os


def extract_infos_from_pdf(pdf_file):
    doc = fitz.open(pdf_file)
    infos = []

    for page_num in range(len(doc)):
        page = doc.load_page(page_num)
        text = page.get_text()

        pattern_emails = re.findall(
            r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', text)
        pattern_jumins = re.findall(r'\d{6}-\d{7}', text)
        pattern_credit_num = re.findall(
            r'\b\d{4}-\d{4}-\d{4}-\d{4}\b', text)
        pattern_cellphone_num = re.findall(
            r'(?:(010-\d{4})|(01[16789]-\d{3,4}))-(\d{4})', text)
        pattern_driver = re.findall(r'\d{2}-\d{2}-\d{6}-\d{2}', text)
        pattern_passport = re.findall(r'([a-zA-Z]{1}|[a-zA-Z]{2})\d{8}', text)
        pattern_account = re.findall(
            r'^(\d{1,})(-(\d{1,})){1,}', text)
        pattern_health = re.findall(r'[1257][-~.\s][0-9]{10}', text)
        # pattern_foreign = re.findall(r'([01][0-9]{5}[\s~-]+[1-8][0-9]{6}|[2-9][0-9]{5}[\s~-]+[1256][0-9]{6})', text)

        for email in pattern_emails:
            infos.append((os.path.basename(pdf_file),
                          page_num + 1, '이메일', email))
        for jumin in pattern_jumins:
            infos.append((os.path.basename(pdf_file),
                          page_num + 1, '주민등록번호', jumin))
        for credit in pattern_credit_num:
            infos.append((os.path.basename(pdf_file),
                          page_num + 1, '신용카드번호', credit))
        for cellphone in pattern_cellphone_num:
            infos.append((os.path.basename(pdf_file),
                          page_num + 1, '휴대전화번호', cellphone))
        for driver in pattern_driver:
            infos.append((os.path.basename(pdf_file),
                         page_num+1, '운전면허번호', driver))
        for passport in pattern_passport:
            infos.append((os.path.basename(pdf_file),
                         page_num + 1, '여권번호', passport))
        for account in pattern_account:
            infos.append((os.path.basename(pdf_file),
                         page_num + 1, '계좌번호', account))
        for health in pattern_health:
            infos.append((os.path.basename(pdf_file),
                         page_num + 1, '건강보험번호', health))
        # for foreign in pattern_foreign:
        #     infos.append((os.path.basename(pdf_file),
        #                  page_num + 1, '외국인등록번호', foreign))

    return infos


def save_infos_to_excel(infos, excel_file):
    if os.path.exists(excel_file):
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["연번", "파일명", "페이지번호", "유형", "내용"])
    start_no = sheet.max_row if sheet.cell(
        row=1, column=1).value == "No." else 0

    for i, info in enumerate(infos, start=start_no + 1):
        sheet.append([i] + list(info))

    workbook.save(excel_file)


def processing_folder(folder_path, excel_file):
    infos_list = []

    for root, _, files in os.walk(folder_path):
        for filename in files:
            if filename.lower().endswith('.pdf'):
                pdf_file_path = os.path.join(root, filename)
                infos = extract_infos_from_pdf(pdf_file_path)
                infos_list.extend(infos)

    save_infos_to_excel(infos_list, excel_file)


if __name__ == "__main__":
    folder_path = input("PDF파일이 존재하는 폴더 경로를 입력하세요: ")
    excel_file = input(
        "엑셀파일 경로를 입력하세요(확장자포함. 파일이 존재하지 않을 경우 새로 생성): ")
    processing_folder(folder_path, excel_file)
    print(f"{excel_file}에 개인정보목록이 생성되었습니다.")
