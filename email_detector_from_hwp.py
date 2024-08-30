import re
import openpyxl
import os
import win32com.client as win32


def extract_infos_from_hwp(hwp_file):
    infos = []
    hwp = None
    try:
        hwp = win32.gencache.EnsureDispatch("HWPFrame.HwpObject")
        hwp.RegisterModule("FilePathCheckDLL", "SecurityModule")
        hwp.Open(hwp_file)
        hwp.InitScan()

        while True:
            state, text = hwp.GetText()
            if state in [0, 1]:
                break
            else:
                result_email = re.search(
                    r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', text)
                if result_email:
                    hwp.MovePos(201)
                    infos.append((os.path.basename(hwp_file),
                                 hwp.KeyIndicator()[3], "이메일",  result_email.group()))
        hwp.ReleaseScan()
        hwp.Quit()

    except Exception as e:
        print(f"한글파일 에러발생 : {e}")
    return infos


def save_infos_to_excel(emails, excel_file):
    if os.path.exists(excel_file):
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["연번", "파일명", "페이지번호", "유형", "내용"])
    start_no = sheet.max_row if sheet.cell(
        row=1, column=1).value == "No." else 0

    for i, email in enumerate(emails, start=start_no + 1):
        sheet.append([i] + list(email))

    workbook.save(excel_file)


def processing_folder(folder_path, excel_file):
    all_emails = []

    for filename in os.listdir(folder_path):
        if filename.lower().endswith('.hwp') or filename.lower().endswith('.hwpx'):
            hwp_file_path = os.path.join(folder_path, filename)
            emails = extract_infos_from_hwp(hwp_file_path)
            all_emails.extend(emails)

    save_infos_to_excel(all_emails, excel_file)


if __name__ == "__main__":
    folder_path = input("개인정보를 추출할 파일이 있는 폴더 경로를 입력하세요: ")
    excel_file = input(
        "엑셀파일 경로를 입력하세요(확장자포함. 파일이 존재하지 않을 경우 새로 생성): ")
    processing_folder(folder_path, excel_file)
    print(f"{excel_file}에 개인정보목록이 생성되었습니다.")
